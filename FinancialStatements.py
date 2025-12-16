import requests
import pandas as pd
import re
from typing import Dict, List, Tuple

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


# Configurations (can be changed based on user preference)
API_KEY = "YOUR_API_KEY_HERE"  # Replace with your Financial Modeling Prep API key (found on their website for free)
TICKER = "V"
BASE = "https://financialmodelingprep.com/stable"

YEARS = 5                 # how many years to pull
SCALE = 1_000             # 1_000 = $000s, 1_000_000 = $MM
OUT_FILE = f"{TICKER}_Financials_Styled.xlsx"


# Creates a prettier label from API generic field names
def prettify_label(s: str) -> str:
    if not isinstance(s, str):
        return str(s)
    s = s.replace("_", " ").replace("-", " ")
    s = re.sub(r"(?<=[a-z])(?=[A-Z])", " ", s)
    s = re.sub(r"(?<=[A-Z])(?=[A-Z][a-z])", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s.title()


def normalize(s: str) -> str:
    return re.sub(r"[^a-z0-9]+", " ", str(s).lower()).strip()


# Fetches the data from the API
def fetch(endpoint: str) -> List[dict]:
    r = requests.get(
        f"{BASE}/{endpoint}",
        params={
            "apikey": API_KEY,
            "symbol": TICKER,
            "period": "annual",
            "limit": YEARS,
        },
        timeout=30,
    )
    if r.status_code == 403: #ususally plan restriction or invalid/blocked key
        raise RuntimeError(
            "403 Forbidden from FMP. Usually: plan restriction, invalid/blocked key. "
            "Print r.text to see the message."
        )
    r.raise_for_status()
    data = r.json()

    if isinstance(data, dict):
        raise RuntimeError(f"FMP error payload: {data}")
    if not isinstance(data, list):
        raise RuntimeError(f"Unexpected response type: {type(data)} -> {data}")

    return data


# Cleans the data and reformats into a dataframe.
def format_statement(data: List[dict], date_col: str = "date", scale: int = SCALE) -> pd.DataFrame:
    df = pd.DataFrame(data).copy()

    if "period" in df.columns:
        df = df[df["period"].astype(str).str.lower().isin(["annual", "fy"])].copy()

    if date_col not in df.columns:
        raise RuntimeError(f"Expected '{date_col}' column in response, got: {df.columns.tolist()}")

    df[date_col] = pd.to_datetime(df[date_col])
    df = df.sort_values(date_col)  # oldest -> newest

    drop_cols = [
        "symbol", "reportedCurrency", "cik", "fillingDate", "filingDate", "acceptedDate",
        "calendarYear", "fiscalYear", "period", "link", "finalLink"
    ]
    df = df.drop(columns=[c for c in drop_cols if c in df.columns], errors="ignore")

    df = df.set_index(date_col).T
    df.columns = [str(c.year) for c in df.columns]
    df = df.reindex(sorted(df.columns), axis=1)

    df.index = [prettify_label(x) for x in df.index]

    num_cols = df.select_dtypes(include="number").columns
    df[num_cols] = df[num_cols] / scale

    return df


# This organizes the rows of the dataframe into buckets based on the provided rules.
def bucket_rows(df: pd.DataFrame, section_rules: List[Tuple[str, List[str]]], other_label: str) -> Dict[str, List[str]]:
    buckets: Dict[str, List[str]] = {sec: [] for sec, _ in section_rules}
    buckets[other_label] = []

    for row in df.index:
        row_n = normalize(row)
        placed = False
        for section, keywords in section_rules:
            for kw in keywords:
                if kw in row_n:
                    buckets[section].append(row)
                    placed = True
                    break
            if placed:
                break
        if not placed:
            buckets[other_label].append(row)

    return buckets

# This builds the ordered list of sections with their corresponding dataframes.
def build_ordered_sections(df: pd.DataFrame, section_rules: List[Tuple[str, List[str]]], other_label: str) -> List[Tuple[str, pd.DataFrame]]:
    buckets = bucket_rows(df, section_rules, other_label)
    sections: List[Tuple[str, pd.DataFrame]] = []

    for section, _ in section_rules:
        rows = buckets.get(section, [])
        if rows:
            sections.append((section, df.loc[rows]))

    other_rows = buckets.get(other_label, [])
    if other_rows:
        sections.append((other_label, df.loc[other_rows]))

    return sections


# Income Statement rules
IS_RULES: List[Tuple[str, List[str]]] = [
    ("Revenue", [
        "revenue", "sales"
    ]),
    ("Cost Of Revenue", [
        "cost of revenue", "cost of goods", "cogs"
    ]),
    ("Gross Profit", [
        "gross profit"
    ]),
    ("Operating Expenses", [
        "selling", "general", "administrative", "sg a", "marketing", "research", "development",
        "operating expense"
    ]),
    ("Operating Income", [
        "operating income", "ebit"
    ]),
    ("Other Income / Expense", [
        "interest", "other income", "other expense", "non operating"
    ]),
    ("Pre-Tax Income", [
        "income before tax", "pretax"
    ]),
    ("Taxes", [
        "tax", "income tax"
    ]),
    ("Net Income", [
        "net income", "net earnings"
    ]),
]

# Balance Sheet rules
BS_RULES: List[Tuple[str, List[str]]] = [
    ("Current Assets", [
        "cash", "receivable", "inventory", "prepaid", "short term", "current asset", "marketable securities"
    ]),
    ("Non-Current Assets", [
        "property", "plant", "equipment", "pp e", "goodwill", "intangible", "long term", "non current asset",
        "deferred tax asset", "right of use", "lease asset", "investment"
    ]),
    ("Current Liabilities", [
        "accounts payable", "payable", "accrued", "current liability", "short term debt",
        "deferred revenue", "tax payable", "current portion"
    ]),
    ("Non-Current Liabilities", [
        "long term debt", "non current liability", "deferred tax liability", "lease liability", "pension"
    ]),
    ("Equity", [
        "equity", "stockholders", "retained earnings", "common stock", "additional paid", "aoci", "treasury"
    ]),
]

# Cash Flow rules
CF_RULES: List[Tuple[str, List[str]]] = [
    ("Cash Flow From Operations", [
        "net income", "depreciation", "amortization", "stock based", "working capital",
        "accounts receivable", "inventory", "accounts payable", "deferred tax", "operating activities"
    ]),
    ("Cash Flow From Investing", [
        "capital expenditure", "capex", "property plant", "acquisition", "investment",
        "investing activities", "purchase of investments", "sale of investments"
    ]),
    ("Cash Flow From Financing", [
        "debt", "repurchase", "dividend", "issuance", "financing activities", "borrow", "repay",
        "common stock", "treasury stock"
    ]),
]


# This provides styling utilities for the Excel sheets.
TITLE_FONT = Font(size=14, bold=True)
HEADER_FONT = Font(bold=True)
SECTION_FONT = Font(bold=True, color="FFFFFF")
SECTION_FILL = PatternFill("solid", fgColor="2F5597")
ALT_ROW_FILL = PatternFill("solid", fgColor="F2F2F2")


def set_col_widths(ws, first_col_width=42, other_col_width=16):
    ws.column_dimensions["A"].width = first_col_width
    for col_idx in range(2, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = other_col_width


def apply_number_formats(ws, start_row: int, start_col: int):
    num_fmt = '#,##0;(#,##0)'
    for r in range(start_row, ws.max_row + 1):
        for c in range(start_col, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            if isinstance(cell.value, (int, float)) and cell.value is not None:
                cell.number_format = num_fmt
                cell.alignment = Alignment(horizontal="right")


def insert_title(ws, title: str, units: str):
    ws.insert_rows(1)
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="left")

    last_col = get_column_letter(ws.max_column)
    ws[f"{last_col}1"] = units
    ws[f"{last_col}1"].font = Font(bold=True, color="666666")
    ws[f"{last_col}1"].alignment = Alignment(horizontal="right")


def freeze_headers(ws):
    ws.freeze_panes = "B3"
    for cell in ws[2]:
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def write_sectioned_sheet(ws, df: pd.DataFrame, sections: List[Tuple[str, pd.DataFrame]]):
    ws.delete_rows(1, ws.max_row)

    # Header row (Row 2)
    ws["A2"] = "Line Item"
    ws["A2"].font = HEADER_FONT
    ws["A2"].alignment = Alignment(horizontal="left")

    years = list(df.columns)
    for j, y in enumerate(years, start=2):
        cell = ws.cell(row=2, column=j, value=y)
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    r = 3

    for section_title, section_df in sections:
        # Section header line
        ws.cell(row=r, column=1, value=section_title).font = SECTION_FONT
        ws.cell(row=r, column=1).fill = SECTION_FILL
        ws.cell(row=r, column=1).alignment = Alignment(horizontal="left")
        for j in range(2, len(years) + 2):
            ws.cell(row=r, column=j, value="").fill = SECTION_FILL

        r += 1

        for row_label in section_df.index:
            ws.cell(row=r, column=1, value=row_label).alignment = Alignment(horizontal="left")

            if (r % 2) == 0:
                for j in range(1, len(years) + 2):
                    ws.cell(row=r, column=j).fill = ALT_ROW_FILL

            for j, y in enumerate(years, start=2):
                val = section_df.loc[row_label, y]
                ws.cell(row=r, column=j, value=None if pd.isna(val) else float(val))

            r += 1

        r += 1


def style_sheet(ws, title: str, units: str):
    insert_title(ws, title, units)
    freeze_headers(ws)
    set_col_widths(ws)
    apply_number_formats(ws, start_row=3, start_col=2)


# Main execution of program
def main():
    income_df = format_statement(fetch("income-statement"))
    bs_df = format_statement(fetch("balance-sheet-statement"))
    cf_df = format_statement(fetch("cash-flow-statement"))

    income_sections = build_ordered_sections(income_df, IS_RULES, other_label="Other Income Statement Items")
    bs_sections = build_ordered_sections(bs_df, BS_RULES, other_label="Other Balance Sheet Items")
    cf_sections = build_ordered_sections(cf_df, CF_RULES, other_label="Other Cash Flow Items")

    # Write initial workbook (sheets will be overwritten with sectioned layout)
    with pd.ExcelWriter(OUT_FILE, engine="openpyxl") as writer:
        income_df.to_excel(writer, sheet_name="Income Statement")
        bs_df.to_excel(writer, sheet_name="Balance Sheet")
        cf_df.to_excel(writer, sheet_name="Cash Flow")

    wb = load_workbook(OUT_FILE)

    units = f"USD (${'000s' if SCALE == 1_000 else 'MM' if SCALE == 1_000_000 else 'scaled'})"

    # Rebuild + style Income Statement
    ws_is = wb["Income Statement"]
    write_sectioned_sheet(ws_is, income_df, income_sections)
    style_sheet(ws_is, f"{TICKER} — Income Statement", units)

    # Rebuild + style Balance Sheet
    ws_bs = wb["Balance Sheet"]
    write_sectioned_sheet(ws_bs, bs_df, bs_sections)
    style_sheet(ws_bs, f"{TICKER} — Balance Sheet", units)

    # Rebuild + style Cash Flow
    ws_cf = wb["Cash Flow"]
    write_sectioned_sheet(ws_cf, cf_df, cf_sections)
    style_sheet(ws_cf, f"{TICKER} — Cash Flow Statement", units)

    wb.save(OUT_FILE)
    print(f"Wrote {OUT_FILE}.")

if __name__ == "__main__":
    main()